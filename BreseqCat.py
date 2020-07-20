#! /usr/bin/python
# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup, UnicodeDammit
import openpyxl
import pandas as pd
import os, argparse, sys
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description="This is a Breseq Mutation Parser.  It Currently only SNPs, Missing Coverage, and New Junction Evidence (wont output junction repeats).  In order to run this program please put all the Breseq directores into a master directory and then parse that directory with the program's -d flag.  Output can be found in BreseqOutput.xlsx")

parser.add_argument('-p', action="store_true", help="Include this flag if you ran population breseq", default=False)
parser.add_argument('-d', action="store", help="Use this flag to indicate the index file you would like to parse", dest ="DirName")

args = parser.parse_args()


if args.DirName == None:
	print("Please include a d flag, try the --help flag if you have questions, exiting!")
	exit(1)
DirName=args.DirName
DirValidation = os.path.isdir(DirName)

if DirValidation not in ['True', True]:
	print("Please Enter a valid Directory to parse, try the --help flag if you have questions, exiting!")
	exit(1)

frames=[]
Total_Line=[]
nje_names=['Sample', 'Seq Id', 'Position', 'Reads (Cov)', 'Reads (Cov)', 'Score', 'Skew', 'Freq', 'Annotation', 'Gene', 'Product']
row_df = pd.DataFrame(columns = nje_names)
One_Total = []
First_One =[]
Last_One = []
Total_List=[]
Total = []
mc_names=['Sample','Seq Id', 'Start', 'End', 'Size', '<-Reads', 'Reads->', 'Gene', 'Description']
ast = "*"
div = "÷"
blank = ""
mc_df = pd.DataFrame(columns = mc_names)
DirectoryNames=os.listdir(DirName)
for i, element in enumerate(DirectoryNames):
	SubDirName=DirectoryNames[i]
	SubDirName=str(SubDirName)

	file = DirName+'/'+SubDirName+'/output/index.html'
	clonal_col_names=['Sample', 'Evidence', 'Seq ID', 'Position', 'Mutation', 'Annotation', 'Gene', 'Description']
	poly_col_names=['Sample', 'Evidence', 'Seq ID', 'Position', 'Mutation', 'Frequency', 'Annotation', 'Gene', 'Description']

	if args.p in ["True", True]:
		snp_df = pd.DataFrame(columns = poly_col_names)
		poly_snp_df = pd.DataFrame(columns = poly_col_names)

	if args.p in ["False", False]:
                snp_df = pd.DataFrame(columns = clonal_col_names)
                poly_snp_df = pd.DataFrame(columns = clonal_col_names)

	with open(file, 'rb') as file:
		soup = BeautifulSoup(file, 'lxml')
		normal_table = soup.find_all(attrs={'class':'normal_table_row'})
		poly_table = soup.find_all(attrs={'class':'polymorphism_table_row'})

		for rows in normal_table:
			row = rows.get_text().encode('utf-8')
			row = row.rstrip('\n')
			line = row[1:].split('\n')
			line = SubDirName.split()+line
			line = tuple(line)
			if args.p in ["True", True]:
				snp_df = snp_df.append(pd.Series(line, index=poly_col_names), ignore_index=True)
			if args.p in ["False", False]:
				snp_df=snp_df = snp_df.append(pd.Series(line, index=clonal_col_names), ignore_index=True)

		for poly_rows in poly_table:
			poly_row = poly_rows.get_text().encode('utf-8')
			poly_row = poly_row.rstrip('\n')
			poly_line = poly_row[1:].split('\n')
			poly_line = SubDirName.split()+poly_line
			poly_line = tuple(poly_line)

                        if args.p in ["True", True]:
                               poly_snp_df = poly_snp_df.append(pd.Series(line, index=poly_col_names), ignore_index=True)
                        if args.p in ["False", False]:
                                poly_snp_df= poly_snp_df.append(pd.Series(line, index=clonal_col_names), ignore_index=True)

		alph_soup = str(soup)
        	begin_umc = alph_soup.find('<tr><th align="left" class="missing_coverage_header_row" colspan="11">Unassigned missing coverage evidence</th></tr>')
        	end_umc = alph_soup.find('<th align="left" class="new_junction_header_row" colspan="12">Unassigned new junction evidence</th>')
        	chopped_soup = alph_soup[begin_umc:end_umc]
        	soup2 = BeautifulSoup(chopped_soup, 'lxml')

        	for rows in soup2:
                	row = rows.get_text().encode('utf-8')
                	row = row.rstrip('\n')
                	line = row[1:].split('\n')

        	line =  line[9:]
        	line = filter(lambda a: a != ast, line)
        	line = filter(lambda a: a != div, line)
        	line = filter(lambda a: a != blank, line)
        	n = len(line)
        	n=n/8

	        for x in range(0,n):
        	        mc_row=line[:8]
			mc_row=SubDirName.split()+mc_row
               		mc_row = tuple(mc_row)
                	line=line[8:]
                	mc_df = mc_df.append(pd.Series(mc_row, index = mc_names), ignore_index = True)

        	alph_soup = str(soup)
        	begin_umc = alph_soup.find('<!-- Item Lines for New Junction -->')
        	chopped_soup = alph_soup[begin_umc:]
        	soup3 = BeautifulSoup(chopped_soup, 'lxml')


        	nje_row_one = soup3.find_all(attrs={'class':'mutation_table_row_0','class':'junction_gene'})
        	nje_repeat = soup3.find_all(attrs={'class':'junction_repeat'})
        	nje_row_center = soup3.find_all(attrs={'align':'center'})
		nje_repeat = str(nje_repeat)

		if len(nje_repeat)==2:

       			for rows in nje_row_center:
                		row = rows.get_text().encode('utf-8')
                		row = row.rstrip('\n')
                		line = row.split('\n')
                		Total_Line=Total_Line+line

        		for rows in nje_row_one:
                		row = rows.get_text().encode('utf-8')
                		row = row.rstrip('\n')
                		line = row.split('\n')
                		One_Total = One_Total+line

			for x in range(0,len(One_Total)/6):
                		First = One_Total[0:1]
                		Last = One_Total[5:6]
                		First_One = First_One + First
                		Last_One = Last_One+ Last
                		One_Total = One_Total[6:]

			for x in range(0,len(Total_Line)/12):
                		Row=Total_Line[:12]
                		Center=Row[2:6]
				FRow=Row[:8]
                		LRow=Row[8:]
                		LSplit=LRow[:2]
                		RSplit=LRow[2:]
                		Last_Row=FRow+LSplit+Center+RSplit
                		NC=First_One[:2]
                		Des=Last_One[:2]
				FRow=Last_Row[:8]
                		LRow=Last_Row[8:]
                		FRow=[SubDirName]+[NC[0]]+FRow+[Des[0]]
                		LRow=[SubDirName]+[NC[1]]+LRow+[Des[1]]
                		FRow=[FRow[0]]+[FRow[1]]+[str("'")+FRow[2]]+FRow[3:]
				FRow=tuple(FRow)
                		LRow=tuple(LRow)
                		row_df = row_df.append(pd.Series(FRow, index=nje_names), ignore_index=True)
                		row_df = row_df.append(pd.Series(LRow, index=nje_names), ignore_index=True)
                		Total_Line=Total_Line[12:]
                		First_One=First_One[2:]
                		Last_One=Last_One[2:]
		if len(nje_repeat)!=2:
			print "Check The error log something happened concatenating " +SubDirName
			text_file=open("Breseq_ConCat_Log.txt", 'a')
			text_file.write("Error Parsing " +SubDirName+"'s breseq, consult the author of the program\n")


	writer = pd.ExcelWriter('Breseq_Output.xlsx', engine='openpyxl')
	frames = frames+[snp_df, poly_snp_df]
	all_snps = pd.concat(frames)
	all_snps.to_excel(writer, 'SNPs', index=False)
	mc_df.to_excel(writer, 'Missing Coverage', index=False)
	row_df.to_excel(writer, 'New Junction Evidence', index=False)
	writer.save()


wb = load_workbook(filename = 'Breseq_Output.xlsx')
ws = wb.get_sheet_by_name(name='New Junction Evidence')

for x in range(0,len(row_df)/2):
	ws.merge_cells('A'+str(2*x+2)+':A'+str(2*x+3))
	ws.merge_cells('E'+str(2*x+2)+':E'+str(2*x+3))
	ws.merge_cells('F'+str(2*x+2)+':F'+str(2*x+3))
	ws.merge_cells('G'+str(2*x+2)+':G'+str(2*x+3))
	ws.merge_cells('H'+str(2*x+2)+':H'+str(2*x+3))
wb.save('Breseq_Output.xlsx')
